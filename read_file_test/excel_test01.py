import pandas as pd
import mysql.connector
from openpyxl import load_workbook


def convert_level(level):
    level_map = {'一级': 1, '二级': 2, '三级': 3, '四级': 4}
    return level_map.get(level, 0)


# 建立到MySQL数据库的连接
conn = mysql.connector.connect(
    host='10.25.10.72',
    port=3306,
    user='root',
    password='000000',
    database='sap'  # 更改为小写的 sap
)
cursor = conn.cursor()

# 读取Excel文件
excel_file = 'Excel01.xlsx'  # 你的Excel文件路径
workbook = load_workbook(excel_file, data_only=True)

# 获取所有sheet页
all_sheets = workbook.sheetnames

# 遍历每个sheet页
for sheet_name in all_sheets:
    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)  # 读取整个sheet页数据
    # 处理连续的空白行情况并跳过空白行
    final_values = []
    current_second_type = None
    df = df[df.iloc[:, 1].notnull()]  # 过滤空白行
    # 处理合并单元格的数据
    df = df.ffill(axis=0)  # 用前面的有效值填充缺失值
    for index, row in df.iterrows():
        value = row[0]
        if pd.isnull(value) or not str(value).strip():
            continue
        else:
            current_second_type = value.strip()
            final_values.append(current_second_type)
    df.iloc[:, 0] = final_values  # 更新DataFrame的第一列数据
    # 处理合并单元格的数据
    merged_rows = []
    current_name = None
    for index, row in df.iterrows():
        name = row[0]
        if pd.isnull(name):
            name = current_name
        else:
            current_name = name
        merged_rows.append(list(row))

    # 将处理后的DataFrame插入到数据库中
    for row_data in merged_rows:
        name = row_data[0]
        wazuh_rule_id = row_data[1]
        wazuh_level = row_data[2]
        level = convert_level(row_data[3])  # 转换中文等级为数字
        description = row_data[4]
        first_type = sheet_name
        second_type = row_data[5]

        # 查询event_category表中对应的id和primary_category
        event_category_id = None
        select_query = """
            SELECT id, primary_category FROM event_category WHERE sub_category = %s
        """
        cursor.execute(select_query, (second_type,))
        result = cursor.fetchone()
        if result:
            event_category_id, primary_category = result

            # 检查一级分类是否匹配，如果不匹配则跳过插入
            if primary_category != first_type:
                continue

        # 插入数据到数据库
        insert_query = """
            INSERT INTO security_event_library_all 
            (name, wazuh_rule_id, wazuh_level, level, description, first_type, second_type, event_category_id) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        values = (name, wazuh_rule_id, wazuh_level, level, description, first_type, second_type, event_category_id)
        cursor.execute(insert_query, values)
        conn.commit()

# 提交并关闭连接
cursor.close()
conn.close()
