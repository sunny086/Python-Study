import pandas as pd
import mysql.connector
from openpyxl import load_workbook


def convert_level(level):
    level_map = {'一级': 1, '二级': 2, '三级': 3, '四级': 4}
    return level_map.get(level, 0)


# 建立到MySQL数据库的连接
conn = mysql.connector.connect(
    host='10.25.10.72',
    port=3306,
    user='root',
    password='000000',
    database='sap'  # 更改为小写的 sap
)
cursor = conn.cursor()

# 读取Excel文件
excel_file = 'Excel01.xlsx'  # 你的Excel文件路径
workbook = load_workbook(excel_file, data_only=True)

# 获取所有sheet页
all_sheets = workbook.sheetnames

# 遍历每个sheet页
for sheet_name in all_sheets:
    df = pd.read_excel(excel_file, sheet_name=sheet_name)

    # 将合并单元格拆分为独立的行
    rows = []
    current_name = None
    for index, row in df.iterrows():
        if pd.notnull(row[0]):
            current_name = row[0]
        row = row.tolist()
        row[0] = current_name
        rows.append(row)

    # 将处理后的数据插入到数据库中
    for row_data in rows:
        name = row_data[0]
        wazuh_rule_id = row_data[1]
        wazuh_level = row_data[2]
        level = convert_level(row_data[3])  # 转换中文等级为数字
        description = row_data[4]
        first_type = sheet_name
        second_type = row_data[5]

        # 查询event_category表中对应的id和primary_category
        event_category_id = None
        select_query = """
            SELECT id, primary_category FROM event_category WHERE sub_category = %s
        """
        cursor.execute(select_query, (second_type,))
        result = cursor.fetchone()
        if result:
            event_category_id, primary_category = result

            # 检查一级分类是否匹配，如果不匹配则跳过插入
            if primary_category != first_type:
                continue

        # 插入数据到数据库
        insert_query = """
            INSERT INTO security_event_library_all 
            (name, wazuh_rule_id, wazuh_level, level, description, first_type, second_type, event_category_id) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        values = (name, wazuh_rule_id, wazuh_level, level, description, first_type, second_type, event_category_id)
        cursor.execute(insert_query, values)

# 提交并关闭连接
conn.commit()
cursor.close()
conn.close()
